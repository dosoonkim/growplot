from scipy.optimize import curve_fit
from math import log
from scipy import stats
import numpy as np
import seaborn as sns
import pandas as pd
import scipy.signal
import sys
from datetime import datetime
from datetime import date
import os
import argparse
from tqdm import tqdm

import matplotlib.pyplot as plt
plt.rcParams['font.sans-serif'] = ['Arial']
plt.rcParams['svg.fonttype'] = 'none'
import json
import jsonpickle
import copy
import scipy.stats


def get_hrs(time_str):
    try:
        h, m, s = time_str.split(':')
    except:
        print(time_str)
        raise ValueError
    return ((int(h) * 3600 + int(m) * 60 + int(s))/3600)


def f(x, L, k, x0):
    return L/(1+np.exp(-k*(x-x0)))

def extend_if_present(map, key, val):
    # We have a lot of dicts where a value is a list, and we need
    # a semantic where if a key already exists in a dict, we extend
    # the list.
    if key in map:
        map[key].extend(val)
    else:
        map[key] = val

def flatten(l_of_l):
    l = []
    for l_ in l_of_l:
        if type(l_) is list:
            l.extend(l_)
        else:
            l.append(l_)
    return l

def av(l):
    return sum(l)/len(l)


# AMW: now calculating SDs as SD of separate parameter estimates.

#font = {'family' : 'normal',
#    'weight' : 'bold',
#    'size'   : 24}

#matplotlib.rc('font', **font)

# Which cells contain the data?
# 1. Search column B for the word 'Time'
# 2. Rows where column B == Time have the headers in them
# 3. Each data series goes down from there until hitting a blank.

from openpyxl import load_workbook
def process_excel(samples, filenames):
    """
    Data will now be keyed on (filename, replicate name) -- which is
    appropriate because that's how we changed the data taken from
    replicates!
    """

    def validate_data_column(filename, data_header, samples):
        """
        Returns true -- that this is a valid data column to be reading --
        if the layout sheet included this column's header, for this filename.
        """
        if data_header == 'Time': return True

        valid = False
        for sample in samples:
            if (filename, data_header) in sample["iSAT_series_IDs"]:
                valid = True
                break
        return valid

    def process_possible_datetime(t):
        v = None
        if type(t) is datetime:
            v = datetime.combine(date.today(), t.time())
        else:
            v = datetime.combine(date.today(), t)
        return v

    def label(letter, val):
        """
        Utility function to get a label string
        """
        return "{}{}".format(letter, val)

    # point of now passing "samples" is so we can exclude irrelevant series early on.
    # like the temperature, which gets misread

    data = {}
    final_times = []

    print("Processing data from {} files...".format(len(filenames)))
    for filename in tqdm(filenames):
        workbook = load_workbook(filename=filename)

        # Workbook should have at least two sheets
        assert(len(workbook.sheetnames) >= 2)

        # Select the non layout sheet.
        sheet = None
        for name in workbook.sheetnames:
            if name != "Layout":
                sheet = workbook[name] # the other one

        times = []

        # Find all the header rows (they're in all the same rows as 'Time').
        header_rows = []
        for ii in range(1,1001):
            if sheet["B{}".format(ii)].value == 'Time': header_rows.append(ii)

        for hr in header_rows:
            for cell in sheet[hr]:
                if not validate_data_column(filename, cell.value, samples): continue

                # Probably shouldn't have passed validation, but definitely not valid!
                if cell.value == "" or cell.value is None: continue

                # Most of the time, we get some y-axis data.
                if cell.value != "Time":
                    data[(filename, cell.value)] = []
                    inc = 1
                    new_label = label(cell.column_letter, hr+inc)
                    while sheet[new_label].value is not None and sheet[new_label].value != "":
                        data[(filename, cell.value)].append(float(sheet[new_label].value))
                        inc += 1
                        new_label = label(cell.column_letter, hr+inc)

                # Sometimes, we get our x-axis
                else:
                    inc = 1
                    new_label = label(cell.column_letter, hr+inc)
                    orig_new_label = new_label
                    while sheet[new_label].value not in [None, ""]:
                        v = process_possible_datetime(sheet[new_label].value)

                        try:
                            times.append( abs(v - sheet[orig_new_label].value).total_seconds()/3600 )
                        except TypeError:
                            raise TypeError

                        inc += 1
                        new_label = label(cell.column_letter, hr+inc)

        # We assume here that all the times are alike. We could implement more stringent checks.
        newtimes = sorted(list(set([t-times[0] for t in times])))
        if len(final_times) == 0: final_times = newtimes
        else:
            try:
                assert(newtimes[-1] == 20)
            except:
                print(filename)
        try:
            if final_times[-1] > 20 and final_times[0] + 20 == final_times[-1]:
                # just decrement
                final_times = [t-final_times[0] for t in final_times]
        except:
            print(filename)

    print("Obtained", len(data.items()), "distinct data series.")
    return data, final_times


def process_labeled_series_group(label, xdata, ydata_series):
    """
    Go through a group of replicates for the same condition, and run curve fits
    on each one. Then gather statistics on the results of those fits.
    """

    paramses = []
    pcoves = []
    for ydata in ydata_series:
        try:
            params, pcov = curve_fit(f, xdata, ydata, maxfev=800000, p0=[max(ydata), 0.1, 6])
        except:
            try:
                params, pcov = curve_fit(f, xdata, ydata, maxfev=80000000, p0=[max(ydata), 0.1, 6])
            except:
                print("issue with {} series".format(label))

        residuals = ydata - f(xdata, *params)
        ss_res = np.sum(residuals**2)
        ss_tot = np.sum((ydata-np.mean(ydata))**2)
        r_squared = 1 - (ss_res / ss_tot)

        paramses.append(params)
        pcoves.append(pcov)

    # Obtain average and standard deviations for each parameter.
    params = [av([param[i] for param in paramses]) for i in range(len(paramses[0]))]
    stdevs = [np.std([param[i] for param in paramses]) for i in range(len(paramses[0]))]

    # The following calculations omit covariance among parameters, hoping it will
    # be close enough to zero.
    half_max = params[0]/2
    half_max_sd = stdevs[0]/2

    max_slope = (params[0]*params[1])/4
    max_slope_sd = ( (stdevs[0]**2 * stdevs[1]**2 + stdevs[0]**2 * params[0]**2 + stdevs[1]**2 * params[1]**2  )**0.5 ) / 4

    half_max_over_max_slope = half_max / max_slope
    # minimal consideration of covariance would append -2cov(x,y)/xy to the second term
    half_max_over_max_slope_sd = ( half_max**2 / max_slope**2 * ( half_max_sd**2 / half_max**2 + max_slope_sd**2/max_slope**2 ) )**0.5

    lag_time = params[2] - half_max_over_max_slope
    # minimal consideration of covariance would subtract -2cov(x,y) inside the sqrt
    lag_time_sd = ( stdevs[2]**2 + half_max_over_max_slope_sd**2 ) ** 0.5

    # first index where ydata is 20x above its baseline (max :6 - min :6) value.
    alternative_lag = 0
    yds = [ys for ys in ydata_series]
    initial_vals = []
    for yd in yds: initial_vals.append(yd[:6])
    baseline = baseline_of(flatten(initial_vals))
    for idx in range(6, len(yds[0])):
        if av([yd[idx] for yd in yds]) > baseline:
            alternative_lag = idx*5
            break
    if alternative_lag > 150 or alternative_lag == 0: alternative_lag = "NA"

    return {'tag': label, 'paramses': paramses, 'params': params, 'stdevs': stdevs, 'max': params[0], 'max_sd': stdevs[0], 'max_slope': max_slope, 'max_slope_sd': max_slope_sd, 'lag_time': lag_time, 'lag_time_sd': lag_time_sd, 'alternative_lag': alternative_lag, 'rsq': r_squared}


def read_layout(filenames, sucrose=False):
    """
    Suppose you have multiple files and you need to aggregate layouts
    together. We need to make sure that stuff gets mapped together!
    Series - magnesium - temperature is the "key" here. Since replicates
    are uniquely named within a file, we have to pair the filename with the
    replicate now.
    """

    print("Reading in {} files...".format(len(filenames)))
    cumulative_samples = []
    tracking_set = set()
    for filename in tqdm(filenames):
        df_samples = pd.read_excel(filename, sheet_name='Layout')
        fn_samples = []
        for _, series in df_samples.iterrows():
            if (filename, series['Design'], series['Mg_Conc (mM)'], series['Temperature']) in tracking_set:
                # careful "append behavior"
                for sample in fn_samples:
                    if sample['series_tag'] != series['Design']: continue
                    if sample['magnesium'] != series['Mg_Conc (mM)']: continue
                    if sample['temperature'] != series['Temperature']: continue
                    if sample['experiment'] != filename: continue
                    sample['iSAT_series_IDs'].extend([(filename, rep.strip().upper()) for rep in series['Replicates'].replace(', ', ',').split(',')])
            else:
                if sucrose:
                    fn_samples.append({
                        'series_tag': series['Design'],
                        'magnesium': series['Mg_Conc (mM)'],
                        'temperature': series['Temperature'],
                        'experiment': filename,
                        'iSAT_series_IDs': [(filename, rep.strip().upper()) for rep in series['Replicates'].replace(', ', ',').split(',')],
                        'fractionator_filename': series['Fractionator']
                    })
                else:
                    fn_samples.append({
                        'series_tag': series['Design'],
                        'magnesium': series['Mg_Conc (mM)'],
                        'temperature': series['Temperature'],
                        'experiment': filename,
                        'iSAT_series_IDs': [(filename, rep.strip().upper()) for rep in series['Replicates'].replace(', ', ',').split(',')],
                    })
                tracking_set.add((filename, series['Design'], series['Mg_Conc (mM)'], series['Temperature']))

        cumulative_samples.extend(fn_samples)

    def reference_conditions_match(ref, sample):
        return sample['magnesium'] == ref['magnesium'] and sample['temperature'] == ref['temperature'] and sample['experiment'] == ref['experiment']

    for ref in tqdm(cumulative_samples):
        if ref['series_tag'] in ['WT_Pilot', 'WT_R1']:
            key = 'normalize_by'
        elif ref['series_tag'] == 'Blank':
            key = 'blank'
        else: continue

        for sample in cumulative_samples:
            if reference_conditions_match(ref, sample):
                extend_if_present(sample, key, ref['iSAT_series_IDs'])

    for sample in cumulative_samples:
        sample['normalize_by'] = list(set(sample['normalize_by']))
        sample['blank'] = list(set(sample['blank']))

    return cumulative_samples


def series_avg(data_serieses):
    return [np.mean(it) for it in zip(*data_serieses)]

def background_subtract(samples, data):
    """
    subtract the corresponding blank for every sample.
    """

    # construct map from each sample tag => the blanks
    # to be subtracted.

    tag_to_blank = {}
    for sample in samples:
        for k in sample['iSAT_series_IDs']:
            tag_to_blank[k] = sample['blank']

    to_del = []
    data_copy = copy.deepcopy(data)
    for k, v in data.items():
        blank_tags = []
        try:
            blank_tags = tag_to_blank[k]
        except KeyError:
            # Only trigger for this should be that we are looking for a series like the temperature,
            # which does not have a blank. Good! We can clean it out here.
            #print(k)
            to_del.append(k)

        data_copy[k] = [v_ - bg_ for (v_, bg_) in zip(v, series_avg([data[blank] for blank in blank_tags ] ))]

    for k in to_del:
        del(data_copy[k])

    return copy.deepcopy(data_copy)


def normalize_to_WT(samples, data):
    """
    divide by the corresponding WT for every sample.
    """

    # construct map from each sample tag => the blanks
    # to be subtracted.

    tag_to_norm = {}
    for sample in samples:
        for k in sample['iSAT_series_IDs']:
            tag_to_norm[k] = sample['normalize_by']

    mg, temp = {}, {}
    for sample in samples:
        for k in sample['iSAT_series_IDs']:
            mg[k] = sample['magnesium']
            temp[k] = sample['temperature']

    to_del = []
    data_copy = copy.deepcopy(data)
    print("Normalizing {} series...".format(len(data.items())))
    for k, v in tqdm(data.items()):
        norm_tags = []
        try:
            norm_tags = tag_to_norm[k]
        except KeyError:
            # Only trigger for this should be that we are looking for a series like the temperature,
            # which does not have a blank. Good! We can clean it out here.
            #print(k)
            to_del.append(k)

        #data[k] = [v_ / bg_ for (v_, bg_) in zip(v, series_avg([data[WT] for WT in norm_tags ] ))]
        coeff = max(series_avg([data[WT] for WT in norm_tags ] ))
        if mg[k] == 3.75 and temp[k] == 37:
            data_copy[k] = [0.43*v_ / coeff for v_ in v]
            # print("LOW MG ax {}. Max of the corresponding WTs is {} which corresponds to a simulated 'normal WT' of {}.".format(max(v), coeff, coeff/0.43))
        elif mg[k] == 3.75 and temp[k] == 30:
            data_copy[k] = [0.88*v_ / coeff  for v_ in v]
            # print("LOW MG LOW TEMP {}. Max of the corresponding WTs is {} which corresponds to a simulated 'normal WT' of {}.".format(max(v), coeff, coeff/0.88))
        elif mg[k] == 7.5 and temp[k] == 37:
            data_copy[k] = [v_ / coeff for v_ in v]
            # print("NORMAL Max {}. Max of the corresponding WTs is {} which corresponds to a simulated 'normal WT' of {}.".format(max(v), coeff, coeff))
        else:
            print("sample at mg {} and temp {} for which we don't have a normalization factor yet!".format(sample['magnesium'], sample['temperature']))
            quit()

    for k in to_del:
        del(data_copy[k])

    return copy.deepcopy(data_copy)


class PlotData:
    def __init__(self, j=None):
        if j is not None:
            self.__dict__ = json.loads(j)
        else:
            self.data = {}
            self.sucrose = []
        #self.label = ""

    def toJSON(self):
        return json.dumps(self, default=lambda o: o.__dict__)


def obtain_parameters(samples, isat_files):
    #df_dict, t = process_excel(filename=isat_file)
    data, time = process_excel(samples=samples, filenames=isat_files)

    #print(time)
    data = background_subtract(samples, data)
    data = normalize_to_WT(samples, data)

    plottable_isat_data = {}#iSATData()

    print("Processing data from {} samples...".format(len(samples)))
    for sample in tqdm(samples):
        k = sample['series_tag']
        if len(sample['iSAT_series_IDs']) == 0:
            print("Sample {} had no associated well IDs!".format(k))
            continue

        print(data.keys())
        print(sample['iSAT_series_IDs'])
        ydata_series = [data[key] for key in sample['iSAT_series_IDs']]
        xdata = time

        datakey = "{}^^^{}".format(sample['magnesium'], sample['temperature'])

        if k not in plottable_isat_data:
            plottable_isat_data[k] = PlotData()

        if datakey not in plottable_isat_data[k].data:
            plottable_isat_data[k].data[datakey] = { 'xdata': xdata }

        extend_if_present(plottable_isat_data[k].data[datakey], 'keys', sample['iSAT_series_IDs'])
        extend_if_present(plottable_isat_data[k].data[datakey], 'ydata_series', ydata_series)

        # Either fit or re-fit, depending. For the future, we should figure out how to do this only once at the end.
        deets = process_labeled_series_group(
            "{}_{}Mg_{}C".format(k, sample['magnesium'], sample['temperature']),
            xdata,
            plottable_isat_data[k].data[datakey]['ydata_series']
        )

        for k_, v_ in deets.items(): plottable_isat_data[k].data[datakey][k_] = v_

    return plottable_isat_data

def write_params(plottable_isat_data):

    q = open("p_vals.csv", "w")
    paramsout = open("all_params.csv", "w")
    paramsout.write("{},{},{},{},{},{},{},{},{},{},{},{}\n".format("Label", "Mg", "Temp", "Max", "Max_sd", "Slope", "Slope_sd", "Lag", "Lag_sd", "slope", "intercept", "alternative_lag"))
    wt_written_params = False
    for k, v in tqdm(plottable_isat_data.items()):
        for k2, v2 in sorted(v.data.items()):
            corresponding_wt_key = 'WT_Pilot' # it's always WT, and the temp and Mg match. '_'.join([k.split('_')[0], 'WT'])
            if 'R1' in k: corresponding_wt_key = 'WT_R1'

            paramses = v2['paramses']

            L  = plottable_isat_data[k].data[k2]['params'][0]
            s  = plottable_isat_data[k].data[k2]['params'][1]
            x0 = plottable_isat_data[k].data[k2]['params'][2]

            if k not in ['WT_Pilot', 'WT_R1']:

                WT_paramses = plottable_isat_data[corresponding_wt_key].data[k2]['paramses']

                _, p_val = scipy.stats.ttest_ind(np.array([p[0] for p in paramses]), np.array([WTp[0] for WTp in WT_paramses]), equal_var=False)
                q.write("{},{},{},{:2.3E}\n".format(k, k2.split("^^^")[0], k2.split("^^^")[1], p_val))
                paramsout.write("{},{},{},{},{},{},{},{},{},{},{},{}\n".format(k, k2.split("^^^")[0], k2.split("^^^")[1],
                    plottable_isat_data[k].data[k2]['max'],
                    plottable_isat_data[k].data[k2]['max_sd'],
                    plottable_isat_data[k].data[k2]['max_slope'],
                    plottable_isat_data[k].data[k2]['max_slope_sd'],
                    plottable_isat_data[k].data[k2]['lag_time'],
                    plottable_isat_data[k].data[k2]['lag_time_sd'],
                    L*s/4,
                    L/2-L*s/4*x0,
                    plottable_isat_data[k].data[k2]['alternative_lag']))
            elif wt_written_params == False:
                paramsout.write("{},{},{},{},{},{},{},{},{},{},{},{}\n".format(k, k2.split("^^^")[0], k2.split("^^^")[1],
                    plottable_isat_data[k].data[k2]['max'],
                    plottable_isat_data[k].data[k2]['max_sd'],
                    plottable_isat_data[k].data[k2]['max_slope'],
                    plottable_isat_data[k].data[k2]['max_slope_sd'],
                    plottable_isat_data[k].data[k2]['lag_time'],
                    plottable_isat_data[k].data[k2]['lag_time_sd'],
                    L*s/4,
                    L/2-L*s/4*x0,
                    plottable_isat_data[k].data[k2]['alternative_lag']))


def make_final_plots(plottable_isat_data, sucrose=False, only_fits=False):

    for k, v in tqdm(plottable_isat_data.items()):
        # set up grid plot?
        # If sucrose, 4; otherwise 3
        num = 3
        if sucrose and len(v.sucrose[0]) != 0:
            num = 4
        fig, axes = plt.subplots(num, 1, gridspec_kw = {'height_ratios':[1]*num}, constrained_layout=True)

        def set_size(w,h, ax=None):
            """ w, h: width, height in inches """
            if not ax: ax=plt.gca()
            l = ax.figure.subplotpars.left
            r = ax.figure.subplotpars.right
            t = ax.figure.subplotpars.top
            b = ax.figure.subplotpars.bottom
            figw = float(w)/(r-l)
            figh = float(h)/(t-b)
            ax.figure.set_size_inches(figw, figh)
        
        set_size(4, 4*num)
        print("Setting size to 4x{} because {}".format(4*num, k))
        
        axes = flatten(axes)
        for ax, (k2, v2) in zip(axes[:len(v.data.items())+1], sorted(v.data.items())):
            corresponding_wt_key = 'WT_Pilot' # it's always WT, and the temp and Mg match. '_'.join([k.split('_')[0], 'WT'])
            if 'R1' in k: corresponding_wt_key = 'WT_R1'

            xdata = v2['xdata']
            ydata_series = v2['ydata_series']
            params = v2['params']
            paramses = v2['paramses']
            stdevs = v2['stdevs']

            yav = series_avg(ydata_series)
            ydata = flatten(ydata_series)

            # normalized
            ax.set_ylim([0,1.5])
            ax.set_yticks([0,0.25,0.5,0.75,1.0,1.25,1.25])

            if not only_fits:
                #ax.scatter(xdata, yav, alpha=0.8, s=4, c='black', label=k)
                ax.scatter([xdata]*(len(ydata)//len(yav)), ydata, alpha=0.3, s=1)

            line_x = np.linspace(0,20,100)
            L  = plottable_isat_data[k].data[k2]['params'][0]
            s  = plottable_isat_data[k].data[k2]['params'][1]
            x0 = plottable_isat_data[k].data[k2]['params'][2]

            line_y = L*s/4* (line_x-x0) + L/2

            ax.plot(line_x,line_y)

            xdp = sorted(xdata)
            reg_curve, = ax.plot(xdp, f(xdp, *params), label="{}_fit".format(k))

            if k != corresponding_wt_key:

                WT_xdata = plottable_isat_data[corresponding_wt_key].data[k2]['xdata']
                WT_params = plottable_isat_data[corresponding_wt_key].data[k2]['params']
                WT_paramses = plottable_isat_data[corresponding_wt_key].data[k2]['paramses']

                ax.plot(sorted(WT_xdata), f(sorted(WT_xdata), *WT_params), label="{}_fit".format(corresponding_wt_key))

                # Where should text be? x is fixed, y depends on design success basically
                _, p_val = scipy.stats.ttest_ind(np.array([p[0] for p in paramses]), np.array([WTp[0] for WTp in WT_paramses]), equal_var=False)
                text_y_loc = 0.15
                if max(ydata) < 0.3: text_y_loc = 1.25
                ax.text(10, text_y_loc, 'p = {:2.2E}'.format(p_val))

            grid = np.linspace(min(xdata), max(xdata), 100)

            # error bands is EITHER a CI for each y value OR it's upper values for each x and then lowers for each x.
            # generate
            n_stdevs = 1.96
            eight_params_sets = [
                    [params[0]-n_stdevs*stdevs[0], params[1]-n_stdevs*stdevs[1], params[2]-n_stdevs*stdevs[2]],
                    [params[0]-n_stdevs*stdevs[0], params[1]-n_stdevs*stdevs[1], params[2]+n_stdevs*stdevs[2]],
                    [params[0]-n_stdevs*stdevs[0], params[1]+n_stdevs*stdevs[1], params[2]-n_stdevs*stdevs[2]],
                    [params[0]-n_stdevs*stdevs[0], params[1]+n_stdevs*stdevs[1], params[2]+n_stdevs*stdevs[2]],
                    [params[0]+n_stdevs*stdevs[0], params[1]-n_stdevs*stdevs[1], params[2]-n_stdevs*stdevs[2]],
                    [params[0]+n_stdevs*stdevs[0], params[1]-n_stdevs*stdevs[1], params[2]+n_stdevs*stdevs[2]],
                    [params[0]+n_stdevs*stdevs[0], params[1]+n_stdevs*stdevs[1], params[2]-n_stdevs*stdevs[2]],
                    [params[0]+n_stdevs*stdevs[0], params[1]+n_stdevs*stdevs[1], params[2]+n_stdevs*stdevs[2]],
                ]

            yvals = [f(grid, *param_set) for param_set in eight_params_sets]
            min_yval, max_yval = [], []
            for idx in range(len(grid)):
                min_yval.append(min([yval[idx] for yval in yvals]))
                max_yval.append(max([yval[idx] for yval in yvals]))
            err_bands = [min_yval, max_yval]

            #ax = plt.gca()
            ax.fill_between(grid, *err_bands, facecolor=reg_curve.get_color(), alpha=.25)#facecolor=fill_color, alpha=.15)
            #ax.legend()

            ax.set_title(r"vs. WT for {} Mg$^{{2+}}$, {} °C".format(k2.split("^^^")[0], k2.split("^^^")[1]))
            ax.set_xlabel("Hours")
            ax.set_ylabel("sfGFP Fluorescence (normalized)")

        if sucrose and len(v.sucrose[0]) != 0:
            ax = axes[-1]
            transformed_data = [v_ - min(v.sucrose[1]) for v_ in v.sucrose[1]]
            ax.scatter(v.sucrose[0], transformed_data, s=1)

            # first_minimum = v.sucrose[1].index(min(v.sucrose[1][300:800]))
            first_minimum = transformed_data.index(min(transformed_data[94:180]))
            ax.set_ylim([0, 1.2*max(transformed_data[first_minimum:])])

            dx = float(max(v.sucrose[0])-min(v.sucrose[0]))/len(v.sucrose[0])
            sigma = 2
            gx = np.arange(-3*sigma, 3*sigma, dx)
            gaussian = np.exp(-(gx/sigma)**2/2)
            result = np.convolve(transformed_data, gaussian, mode="full")

            #print(scipy.signal.find_peaks(v.sucrose[1]))
            minima = scipy.signal.find_peaks(-1*result)[0]
            print(k, minima)
            if len(minima) < 4: ax.text(30, 6*max(transformed_data[first_minimum:]), 'Ratio: 0')
            else:
                # integrate 0 to 1, 1 to 2, compare to 2 to 3
                # don't have to multiply because the incorrect
                # delta X will cancel out in the ratio
                # integrated_30S = sum(transformed_data[minima[0]:minima[1]])
                # integrated_50S = sum(transformed_data[minima[1]:minima[2]])
                # integrated_70S = sum(transformed_data[minima[2]:minima[3]])
                # min closest to 550 to min closest to
                integrated_50S = sum(transformed_data[minima[2]:minima[3]])
                integrated_70S = sum(transformed_data[minima[3]:])
                # ax.text(30, 0.6*max(transformed_data[first_minimum:]), 'Ratio: {:2.2f}'.format(integrated_70S/(integrated_50S+integrated_30S)))
                ax.text(30, 0.6*max(transformed_data[first_minimum:]), 'Ratio: {:2.2f}'.format(integrated_70S/(integrated_50S)))

        #plt.constrained_layout()

        fig.suptitle('Design {}'.format(k), fontsize=16)

        plt.savefig('{}.png'.format(k), dpi=600)
        #plt.rcParams['svg.fonttype']='none'
        #plt.savefig('plot.svg')
        plt.clf()
        plt.close()

def read_sucrose_file(fn):
    if type(fn) is float: # cheap isnan
        return ([], [])
    if not os.path.exists("sucrose/{}".format(fn)):
        return ([], [])

    x, y = [], []
    with open("sucrose/{}".format(fn)) as f:
        print("Reading", fn)
        header_found = False
        for line in f:
            if 'AbsA' in line:
                header_found = True
                continue
            if header_found:
                line = line.split(',')
                x_ = float(line[4])
                y_ = float(line[5])
                x.append(x_)
                y.append(y_)
    return x,y

def obtain_sucrose_data(samples, plot_data):
    for sample in samples:
        plot_data[sample['series_tag']].sucrose = read_sucrose_file(sample['fractionator_filename'])

    with open("cache.json", 'w') as f:
        f.write(jsonpickle.encode(plot_data))

def read_cache():
    if os.path.exists("cache.json"):
        with open("cache.json") as f:
            return jsonpickle.decode(f.read())

def baseline_of(arr):
    return 10*(max(arr) - min(arr))

def alternative_lag(ydata):
    # first index where ydata is 20% above its baseline (max :4) value.

    # New calc: you have to exceed the max :4 by 2x max :4 - min :4
    alternative_lag = 0
    baseline = baseline_of(ydata[:6])

    #print("baseline is ", baseline)
    for idx, ydat in enumerate(ydata[6:]):
        if ydat > baseline:
            # print(ydat, baseline, idx+6)
            alternative_lag = (idx+6)*5
            break

    if alternative_lag > 150 or alternative_lag == 0: alternative_lag = "NA"

    return alternative_lag

def write_plot_data(plot_data):
    """ write out plot data for each series
    """

    protodf = []

    for k, v in plot_data.items():
        #Exp	Salts	Temperature	Subunit	Challenge	Design	Sample	Well	t_000
        for k2, v2 in v.data.items():
            for key, series, params_set in zip(v2['keys'], v2['ydata_series'], v2['paramses']):


                L, s, x0  = params_set

                slope = L*s/4
                y0 = L/2 - L*s*x0/4

                datum = {}
                datum['Salts'] = k2.split("^^^")[0]
                datum['Temperature'] = k2.split("^^^")[1]
                datum['Design'] = k
                datum['Well'] = key[1]
                datum['Exp'] = key[0]
                datum['Max_Param'] = L
                datum['Lag_Param'] = x0
                datum['Alternative_Lag'] = alternative_lag(series)
                datum['Slope_Param'] = slope
                datum['Intercept_Param'] = y0
                for xdatum, ydatum in zip(v2['xdata'], series):
                    datum['t_{:03}'.format(int(60*xdatum+0.01))] = ydatum
                protodf.append(datum)

    df = pd.DataFrame.from_dict(protodf)
    df.to_csv("all_data.csv")

if __name__ == '__main__':
    samples = read_layout(["Layout.xlsx"], sucrose=True)
    plot_data = obtain_parameters(samples, ["191203_EteRNA_23S_37C_MgConcs_1.xlsx"] )
    obtain_sucrose_data(samples, plot_data)
    make_final_plots(plot_data, sucrose=True)




